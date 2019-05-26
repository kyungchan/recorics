from openpyxl import load_workbook
from collections import OrderedDict
import argparse
import sys
import operator
import os
import json

parser = argparse.ArgumentParser()
parser.add_argument(
    'data', help='lyric data you wanna compare with specific sentence you will provide')
args = parser.parse_args()

data = args.data
base = data

wb = load_workbook(os.path.abspath(
    './public/python_module/lyric_similarity') + '/lyric_data.xlsx')
track_sheet = wb.worksheets[0]
sentence_sheet = wb.worksheets[1]
max_row = sentence_sheet.max_row

sentence_list = []
track_id_list = []

for i in range(2, max_row+1):
    sentence_list.append(sentence_sheet['B%d' % i].value)
    track_id_list.append(sentence_sheet['A%d' % i].value)

max_row = track_sheet.max_row

track_song_info_dict = {}
track_artist_info_dict = {}
track_num_line_info_dict = {}

for i in range(2, max_row+1):
    track_song_info_dict[track_sheet['C%d' %
                                     i].value] = track_sheet['B%d' % i].value
    track_artist_info_dict[track_sheet['C%d' %
                                       i].value] = track_sheet['A%d' % i].value
    track_num_line_info_dict[track_sheet['C%d' %
                                         i].value] = track_sheet['D%d' % i].value


# n-gram 유사도 비교


def ngram(s, num):
    res = []
    slen = len(s) - num + 1
    for i in range(slen):
        ss = s[i:i+num]
        res.append(ss)
    return res


def diff_ngram(sa, sb, num):
    a = ngram(sa, num)
    b = ngram(sb, num)
    r = []
    cnt = 0
    for i in a:
        for j in b:
            if i == j:
                cnt += 1
                r.append(i)
    return cnt / len(a), r


three_gram_score_list = []
three_gram_word_list = []

for s in sentence_list:
    # 3-gram
    r3, word3 = diff_ngram(base, s, 3)
    three_gram_score_list.append(r3)
    three_gram_word_list.append(word3)


# 3-gram
three_max_index = three_gram_score_list.index(max(three_gram_score_list))
three_max_track_id = track_id_list[three_max_index]

tmp_sentence_list = []
tmp_track_list = []

max_song_line_number = track_num_line_info_dict[three_max_track_id]
max_song_start_index = track_id_list.index(three_max_track_id)

for i in range(0, max_song_line_number):
    print(sentence_list[max_song_start_index + i] + '/')

print('%s|' % sentence_list[three_max_index])
print('%s|%s|' % (
    track_artist_info_dict[three_max_track_id], track_song_info_dict[three_max_track_id]))

tmp_sentence_list.append(sentence_list[three_max_index])
tmp_track_list.append(three_max_track_id)


# total = 이중리스트
# total = [[아티스트, 곡, 유사도, 문장]]
total = []

try:
    for i in range(0, len(three_gram_score_list)):
        if three_gram_score_list[i] > 0.15:
            if(sentence_list[i] not in tmp_sentence_list and track_id_list[i] not in tmp_track_list):
                tmp_sentence_list.append(sentence_list[i])
                tmp_track_list.append(track_id_list[i])
                total.append([track_artist_info_dict[track_id_list[i]],
                              track_song_info_dict[track_id_list[i]], three_gram_score_list[i], sentence_list[i]])

    total.sort(key=lambda x: x[2], reverse=True)

    if len(total) > 0:
        print(total)
except KeyError:
    sys.exit(1)